# -*- coding: utf-8 -*-
__author__ = 'Kay'


import os
import json
from openpyxl import Workbook


def convert_weibo_xls(in_name, out_name, header):

    book = Workbook()
    sheet = book.active
    sheet.title = 'data'
    for i in list(range(len(header))):
        sheet.cell(row = 1, column = i + 1).value = header[i]

    data = json.load(open(in_name))

    for i, row in enumerate(data):
        sheet.cell(row = i + 2, column = 1).value = row['w:bi_followers_count']
        sheet.cell(row = i + 2, column = 2).value = row['w:gender']
        sheet.cell(row = i + 2, column = 3).value = row['w:ucreated_at']
        sheet.cell(row = i + 2, column = 4).value = row['w:statuses_count']
        sheet.cell(row = i + 2, column = 5).value = row['w:source']
        sheet.cell(row = i + 2, column = 6).value = row['w:followers_count']
        sheet.cell(row = i + 2, column = 7).value = row['w:seg']
        sheet.cell(row = i + 2, column = 8).value = row['w:uid']
        sheet.cell(row = i + 2, column = 9).value = row['w:text']
        sheet.cell(row = i + 2, column = 10).value = row['w:friends_count']
        sheet.cell(row = i + 2, column = 11).value = row['w:created_at']
        sheet.cell(row = i + 2, column = 12).value = row['w:location']
        sheet.cell(row = i + 2, column = 13).value = row['id']
        sheet.cell(row = i + 2, column = 14).value = row['w:screen_name']
        sheet.cell(row = i + 2, column = 15).value = row['w:verified']

    book.save(out_name)


def convert_weibo(in_name, out_name, header):


    out_file = open(out_name, 'w')
    out_file.write('<-->'.join(header) + '\n')

    for d in open(in_name):
        row = json.loads(d)
        line = []
        line.append(row['w:bi_followers_count'])
        line.append(row['w:gender'])
        line.append(row['w:ucreated_at'])
        line.append(row['w:statuses_count'])
        line.append(row['w:source'])
        line.append(row['w:followers_count'])
        line.append(row['w:uid'])
        line.append(row['w:text'])
        line.append(row['w:friends_count'])
        line.append(row['w:created_at'])
        line.append(row['w:location'])
        line.append(row['id'])
        line.append(row['w:screen_name'])
        line.append(row['w:verified'])
        out_file.write('<-->'.join(line) + '\n')

    out_file.close()


if __name__ == '__main__':

    # 微博数据
    header = ['互相关注数', '性别', '用户创建时间', '微博量',
              '来源', '粉丝数', '用户ID',
              '微博内容', '关注数', '微博创建时间', '地理位置',
              '微博ID', '昵称', '是否认证']
    words = ['马化腾', '腾讯', '华为', '任正非']
    for w in words:
        print(w)
        convert_weibo('data/{}.txt'.format(w), '/data2/word/{}.csv'.format(w), header)

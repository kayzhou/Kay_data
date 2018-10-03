# -*- coding: utf-8 -*-
__author__ = 'Kay'


import os
import json
from openpyxl import Workbook


def convert_webpage(in_name, out_name, header):

    book = Workbook()
    sheet = book.active
    sheet.title = 'data'
    for i in list(range(len(header))):
        sheet.cell(row = 1, column = i + 1).value = header[i]

    data = json.load(open(in_name))

    for i, row in enumerate(data):
        try:
            sheet.cell(row = i + 2, column = 1).value = row['type']
            sheet.cell(row = i + 2, column = 2).value = row['url']
            sheet.cell(row = i + 2, column = 3).value = row['website']
            sheet.cell(row = i + 2, column = 4).value = row['title']
            sheet.cell(row = i + 2, column = 5).value = row['author']
            sheet.cell(row = i + 2, column = 6).value = row['content']
            sheet.cell(row = i + 2, column = 7).value = row['created_time']
        except:
            print(row)

    book.save(out_name)


if __name__ == '__main__':

    # 网页数据
    header = ['来源类型', '链接', '网站', '题目',
              '作者', '内容', '创建时间']

    convert_webpage('data/郭广昌.txt', 'data/郭广昌.xlsx', header)
